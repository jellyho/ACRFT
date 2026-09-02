"""Real-robot execution-length sweep on YAM lego-taxi (user's rollouts, LEGOPROG.xlsx).

bcH_ex_K = BC policy trained with horizon H, executing K steps of each predicted chunk
before replanning. progress in 0..4 (assembly stages), 10 rollouts per condition.
Source of record: .scratch/legoprog_v3.xlsx (copied from the user's upload 2026-08-27);
every number is parsed from it here. Run with:  uv run --with openpyxl python scripts/fig_legoprog.py
"""

# ruff: noqa: E402, ICN001  (matplotlib.use must precede pyplot; probe-local imports intentional)

import pathlib
import sys

import matplotlib
import numpy as np

matplotlib.use("Agg")
import matplotlib.pyplot as plt
from openpyxl import load_workbook

R = pathlib.Path(__file__).resolve().parents[1]
sys.path.insert(0, str(R / "slurm"))
import plot_style

plot_style.apply()
PAL = plot_style.PALETTE

ws = load_workbook(R / ".scratch/legoprog_v3.xlsx")[  # one sheet, 8 (name,num,prog) column triplets
    "Sheet1"
]
data = {}
for c in range(1, ws.max_column + 1, 3):
    name = ws.cell(row=2, column=c).value
    vals = [ws.cell(row=r, column=c + 2).value for r in range(2, 12)]
    data[name] = np.array([v for v in vals if v is not None], float)

bc50 = {int(k.split("_")[-1]): v for k, v in data.items() if k.startswith("bc50")}
bc30 = {int(k.split("_")[-1]): v for k, v in data.items() if k.startswith("bc30")}


def mean_ci(v, t95=2.262):  # t(9, .975)
    return v.mean(), t95 * v.std(ddof=1) / np.sqrt(len(v))


fig, ax = plt.subplots(figsize=(4.6, 3.2))
rng = np.random.default_rng(0)
for i, (d, name, marker) in enumerate(((bc50, "trained H=50", "o"), (bc30, "trained H=30", "s"))):
    ks = sorted(d)
    m, ci = zip(*[mean_ci(d[k]) for k in ks], strict=True)
    ax.errorbar(ks, m, yerr=ci, fmt=f"{marker}-", color=PAL[i], lw=1.8, ms=6, capsize=3, label=name)
    for k in ks:  # raw rollouts, jittered
        ax.scatter(
            k + rng.uniform(-1.2, 1.2, len(d[k])),
            d[k] + rng.uniform(-0.07, 0.07, len(d[k])),
            s=9,
            color=PAL[i],
            alpha=0.30,
            linewidths=0,
            zorder=1,
        )
ax.axhline(4.0, color="#555", lw=1.0, ls="--")
ax.text(4.5, 4.08, "complete", ha="left", va="bottom", fontsize=8, color="#555")
ax.set_xlabel("executed steps per chunk")
ax.set_ylabel("progress (0–4)")
ax.set_xticks([5, 10, 20, 30, 40, 50])
ax.set_ylim(-0.25, 4.65)
ax.legend(loc="upper right", bbox_to_anchor=(1.0, 0.97))
fig.tight_layout()
out = R / ".scratch/fig_legoprog"
fig.savefig(f"{out}.png", dpi=220)
fig.savefig(f"{out}.svg")
for k in sorted(bc50):
    m, ci = mean_ci(bc50[k])
    print(f"bc50 ex{k:2d}: {m:.1f} ± {ci:.2f}")
for k in sorted(bc30):
    m, ci = mean_ci(bc30[k])
    print(f"bc30 ex{k:2d}: {m:.1f} ± {ci:.2f}")
print("wrote", out)
