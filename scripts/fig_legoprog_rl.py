"""Real-robot critic-guided deployment results (LEGOPROG v2, lower block).

Base policy = BC h30; fixed_* critics deploy as best-of-N chunk selection, g5_* critics as
adaptive prefix selection. progress 0..4, 10 rollouts per arm. Reference lines = the same
base policy without any critic (bc30_ex_10 / bc30_ex_30 from the upper block).
Source of record: .scratch/legoprog_v2.xlsx. Run: uv run --with openpyxl python scripts/fig_legoprog_rl.py
"""

# ruff: noqa: E402, ICN001  (matplotlib.use must precede pyplot; probe-local imports intentional)

import pathlib
import sys

import matplotlib
import numpy as np

matplotlib.use("Agg")
import matplotlib.pyplot as plt
from openpyxl import load_workbook

R = pathlib.Path(__file__).resolve().parents[1]
sys.path.insert(0, str(R / "slurm"))
import plot_style

plot_style.apply()
PAL = plot_style.PALETTE

ws = load_workbook(R / ".scratch/legoprog_v2.xlsx")["Sheet1"]


def block(header_row):
    out = {}
    for c in range(1, ws.max_column + 1, 3):
        name = ws.cell(row=header_row, column=c).value
        if not name:
            continue
        vals = [ws.cell(row=r, column=c + 2).value for r in range(header_row, header_row + 10)]
        out[str(name)] = np.array([v for v in vals if v is not None], float)
    return out


bc = block(2)
rl = block(14)


def mean_ci(v, t95=2.262):
    return v.mean(), t95 * v.std(ddof=1) / np.sqrt(len(v))


ARMS = [  # (xlsx name, display, family index for color)
    ("fixed", "fixed\n(τ.7)", 0),
    ("fixed_nofloor", "fixed\nnofloor", 0),
    ("fixed_tau9_noaug", "fixed\nτ.9", 0),
    ("fixed_tau9_min", "fixed\nτ.9+aug", 0),
    ("g5", "adaptive\n(τ.7)", 1),
    ("g5_tau9_min", "adaptive\nτ.9+aug", 1),
]

fig, ax = plt.subplots(figsize=(5.4, 3.2))
rng = np.random.default_rng(1)
xs = np.arange(len(ARMS), dtype=float)
xs[4:] += 0.45  # visual gap between BoN (fixed) and adaptive (g5) families
for x, (name, _label, fam) in zip(xs, ARMS, strict=True):
    v = rl[name]
    m, ci = mean_ci(v)
    ax.errorbar([x], [m], yerr=[ci], fmt="o" if fam == 0 else "s", color=PAL[fam], ms=7, capsize=3, lw=1.8, zorder=3)
    ax.scatter(
        x + rng.uniform(-0.14, 0.14, len(v)),
        v + rng.uniform(-0.06, 0.06, len(v)),
        s=10,
        color=PAL[fam],
        alpha=0.3,
        linewidths=0,
        zorder=1,
    )
for ref, style, dy in (("bc30_ex_10", "--", 0.06), ("bc30_ex_30", ":", -0.17)):
    m, _ = mean_ci(bc[ref])
    ax.axhline(m, color="#555", lw=1.0, ls=style)
    ax.text(-0.42, m + dy, ref.replace("bc30_ex_", "BC ex"), fontsize=8, color="#555", ha="left")
ax.set_xticks(xs)
ax.set_xticklabels([a[1] for a in ARMS], fontsize=8.5)
ax.set_ylabel("progress (0–4)")
ax.set_ylim(-0.25, 4.35)
ax.text(np.mean(xs[:4]), 4.15, "best-of-N", ha="center", fontsize=9, color=PAL[0])
ax.text(np.mean(xs[4:]), 4.15, "adaptive", ha="center", fontsize=9, color=PAL[1])
fig.tight_layout()
out = R / ".scratch/fig_legoprog_rl"
fig.savefig(f"{out}.png", dpi=220)
fig.savefig(f"{out}.svg")
for name, _label, _fam in ARMS:
    m, ci = mean_ci(rl[name])
    print(f"{name:18s} {m:.1f} ± {ci:.2f}")
for ref in ("bc30_ex_10", "bc30_ex_30"):
    m, ci = mean_ci(bc[ref])
    print(f"{ref:18s} {m:.1f} ± {ci:.2f} (baseline)")
print("wrote", out)
